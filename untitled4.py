# -*- coding: utf-8 -*-
"""
@author: Erick Gildardo Avalos Soiis
"""
from openpyxl import Workbook 
import openpyxl
import xlrd
from keras.models import Sequential
from keras.layers import Dense
import tensorflow as tf
import numpy as np
columnas= ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
def N_Proyectos():
    book =  openpyxl.load_workbook('Libro.xlsx')
    np = book.sheetnames
    book.close()
    return len(np)

def Proyectos():
    book =  openpyxl.load_workbook('Libro.xlsx')
    np = book.sheetnames
    book.close()
    np.pop(0)
    return np

def p_calcular(N,d,nc):
    book =  openpyxl.load_workbook('Libro.xlsx')
    sheet = book[N]
    nv =(sheet['A2'].value)
    iv =(sheet['A3'].value)
    fv =(sheet['A4'].value)
    pr =(sheet['A5'].value)
    nn =(sheet['A7'].value)
    book.close()
    book =  openpyxl.load_workbook('%s'%d)
    sheet = book.active
    X =[]
    p1=26
    if 26>pr:
        p1=pr
            
    
    for i in range(nc):
        ax=[]
        for j in range(p1):
            ax.append(sheet['%s%d'%(columnas[j],i+1)].value)
        X.append(ax)
        
    if nv>26:
        p2=nv-26
        
        for i in range(nc):
            K = 0
            for j in range(p2):
                X[i].append(sheet['%s%s%d'%(columnas[K],columnas[j],i+1)].value)
                if j >26:
                    K = K +1

    book.close()
    X=tuple(X)
    X=np.array(X)
    return X
    
def p_g_Y(N,d,nc,Y):
    book =  openpyxl.load_workbook('Libro.xlsx')
    sheet = book[N]
    nv =(sheet['A2'].value)
    iv =(sheet['A3'].value)
    fv =(sheet['A4'].value)
    pr =(sheet['A5'].value)
    nn =(sheet['A7'].value)
    book.close()
    
    book =  openpyxl.load_workbook('%s'%d)
    sheet = book.active
    p1=26
    iy = ''
    if 26>pr:
        p1=pr
            
    for i in range (p1):
        iy = columnas[i]
        
    if pr>26:
        p2=nv-26
        K=0
        for j in range(p2):
            iy=('%s%s'%(columnas[K],columnas[j+1]))
            if j >26:
                K = K +1
    for i in range(nc):
        sheet['%s%d'%(iy,(i+1))]=Y[i]
        
    book.save('%s'%d)
    book.close()
    
def retornar_datos(N):
    book =  openpyxl.load_workbook('Libro.xlsx')
    sheet = book[N]
    ret = []
    ret.append(N)
    ret.append(sheet['A1'].value)
    ret.append(sheet['A2'].value)
    ret.append(sheet['A3'].value)
    ret.append(sheet['A4'].value)
    ret.append(sheet['A5'].value)
    ret.append(sheet['A6'].value)
    ret.append(sheet['A7'].value)
    ret.append(sheet['A8'].value)
    book.close()
    return ret
def elimina_hoja(N):
    book =  openpyxl.load_workbook('Libro.xlsx')
    sheet = book[N]
    book.remove(sheet)
    book.save('Libro.xlsx')
    book.close()
    
def crear_txt(n,d,nv,iv,fv,pr,nc,nn,ni):
    book =  openpyxl.load_workbook('Libro.xlsx')
    sheet = book.create_sheet(n)
    sheet['A1']=d
    sheet['A2']=nv
    sheet['A3']=iv
    sheet['A4']=fv
    sheet['A5']=pr
    sheet['A6']=nc
    sheet['A7']=nn
    sheet['A8']=ni
    book.save('Libro.xlsx')
    book.close()
    
    book =  openpyxl.load_workbook('%s'%d)
    sheet = book.active
    X =[]
    p1=26
    Y=[]
    iy = ''
    if 26>pr:
        p1=pr
            
    for i in range (p1):
        iy = columnas[i]
        
    if pr>26:
        p2=nv-26
        K=0
        for j in range(p2):
            iy=('%s%s'%(columnas[K],columnas[j+1]))
            if j >26:
                K = K +1
    for i in range(nc):
        Y.append(sheet['%s%d'%(iy,(i+1))].value)
        
    print(np.asarray(Y))
    
    for i in range(nc):
        ax=[]
        for j in range(p1):
            ax.append(sheet['%s%d'%(columnas[j],i+1)].value)
        X.append(ax)
        
    if nv>26:
        p2=nv-26
        
        for i in range(nc):
            K = 0
            for j in range(p2):
                X[i].append(sheet['%s%s%d'%(columnas[K],columnas[j],i+1)].value)
                if j >26:
                    K = K +1

    book.close()
    Y=np.asarray(Y)
    X=tuple(X)
    X=np.array(X)
    
    return X,Y

    